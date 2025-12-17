import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from tkcalendar import DateEntry
from openpyxl import Workbook
import pandas as pd
import os
from reports import reports_window

DB_NAME = "pharma.db"

# ================== БАЗА ==================

def get_connection():
    return sqlite3.connect(DB_NAME)

def init_db():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        password TEXT
    )""")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS doctors (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        doctor_code TEXT,
        fio TEXT,
        city TEXT,
        region TEXT,
        specialty TEXT,
        coef REAL
    )""")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_code TEXT,
        brand TEXT,
        batch TEXT,
        expiry_date TEXT
    )""")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS sales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        doctor_code TEXT,
        product_code TEXT,
        region TEXT,
        quantity INTEGER
    )""")

    cur.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO users VALUES (NULL,'admin','1234')")

    conn.commit()
    conn.close()

# ================== EXCEL ШАБЛОНЫ ==================

def create_excel_templates():
    os.makedirs("templates", exist_ok=True)

    templates = {
        "doctors_template.xlsx": ["doctor_code","fio","city","region","specialty","coef"],
        "products_template.xlsx": ["product_code","brand","batch","expiry_date"],
        "sales_template.xlsx": ["date","doctor_code","product_code","quantity"]
    }

    for name, headers in templates.items():
        path = os.path.join("templates", name)
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            wb.save(path)

def download_template(name):
    src = os.path.join("templates", name)
    dst = filedialog.asksaveasfilename(
        initialfile=name,
        defaultextension=".xlsx",
        filetypes=[("Excel","*.xlsx")]
    )
    if dst:
        with open(src,"rb") as f1, open(dst,"wb") as f2:
            f2.write(f1.read())
        messagebox.showinfo("OK","Шаблон сохранён")

# ================== ЛОГИН ==================

def login_window():
    root = tk.Tk()
    root.title("Авторизация")
    root.geometry("300x200")

    tk.Label(root,text="Логин").pack()
    e_user = tk.Entry(root); e_user.pack()
    tk.Label(root,text="Пароль").pack()
    e_pass = tk.Entry(root,show="*"); e_pass.pack()

    def login():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username=? AND password=?",
                    (e_user.get(),e_pass.get()))
        if cur.fetchone():
            root.destroy()
            main_menu()
        else:
            messagebox.showerror("Ошибка","Неверный логин")
        conn.close()

    tk.Button(root,text="Войти",command=login).pack(pady=10)
    root.mainloop()

# ================== МЕНЮ ==================

def main_menu():
    root = tk.Tk()
    root.title("Pharma System")
    root.geometry("420x420")

    tk.Label(root,text="Главное меню",font=("Arial",14,"bold")).pack(pady=10)

    tk.Button(root,text="Врачи",width=30,command=doctors_window).pack(pady=3)
    tk.Button(root,text="Препараты",width=30,command=products_window).pack(pady=3)
    tk.Button(root,text="Продажи",width=30,command=sales_window).pack(pady=3)
    tk.Button(root,text="Отчёты",width=30,command=reports_window).pack()
    tk.Button(root,text="лаоылуга",width=30,command=sales_window).pack(pady=3)

    tk.Label(root,text="Сотрудники",font=("Arial",11,"bold")).pack(pady=8)
    tk.Button(root,text="Менеджер",width=30,
              command=lambda: messagebox.showinfo("Роль","Режим: Менеджер")).pack(pady=2)
    tk.Button(root,text="Медпред",width=30,
              command=lambda: messagebox.showinfo("Роль","Режим: Медпред")).pack(pady=2)

    tk.Button(root,text="Выход",width=30,command=root.destroy).pack(pady=10)
    root.mainloop()

# ================== ВРАЧИ ==================

def doctors_window():
    win = tk.Toplevel()
    win.title("Врачи")
    win.geometry("1100x450")

    frame = tk.Frame(win); frame.pack(pady=5)

    labels = ["Код","ФИО","Город","Регион","Специальность","Коэф"]
    entries = []

    for i,l in enumerate(labels):
        tk.Label(frame,text=l).grid(row=0,column=i)
        e = tk.Entry(frame,width=15)
        e.grid(row=1,column=i)
        entries.append(e)

    def load():
        tree.delete(*tree.get_children())
        conn=get_connection();cur=conn.cursor()
        cur.execute("SELECT doctor_code,fio,city,region,specialty,coef FROM doctors")
        for r in cur.fetchall():
            tree.insert("", "end", values=r)
        conn.close()

    def add():
        conn=get_connection();cur=conn.cursor()
        cur.execute("""
            INSERT INTO doctors (doctor_code,fio,city,region,specialty,coef)
            VALUES (?,?,?,?,?,?)
        """, tuple(e.get() for e in entries))
        conn.commit();conn.close();load()

    def delete():
        sel = tree.selection()
        if not sel: return
        code = tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Удалить",f"Удалить врача {code}?"):
            conn=get_connection();cur=conn.cursor()
            cur.execute("DELETE FROM doctors WHERE doctor_code=?", (code,))
            conn.commit();conn.close();load()

    def import_xlsx():
        p=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if not p: return
        df=pd.read_excel(p)
        conn=get_connection();cur=conn.cursor()
        for _,r in df.iterrows():
            cur.execute("""
                INSERT INTO doctors (doctor_code,fio,city,region,specialty,coef)
                VALUES (?,?,?,?,?,?)
            """, (r.get("doctor_code"),r.get("fio"),r.get("city"),
                  r.get("region"),r.get("specialty"),r.get("coef",0)))
        conn.commit();conn.close();load()

    def export_xlsx():
        conn=get_connection();cur=conn.cursor()
        cur.execute("SELECT doctor_code,fio,city,region,specialty,coef FROM doctors")
        rows=cur.fetchall();conn.close()
        os.makedirs("exports",exist_ok=True)
        wb=Workbook();ws=wb.active
        ws.append(["doctor_code","fio","city","region","specialty","coef"])
        for r in rows: ws.append(r)
        wb.save(f"exports/doctors_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

    tk.Button(frame,text="Добавить",command=add).grid(row=1,column=6,padx=3)
    tk.Button(frame,text="Импорт",command=import_xlsx).grid(row=1,column=7,padx=3)
    tk.Button(frame,text="Экспорт",command=export_xlsx).grid(row=1,column=8,padx=3)
    tk.Button(frame,text="Шаблон",
              command=lambda:download_template("doctors_template.xlsx")).grid(row=1,column=9,padx=3)
    tk.Button(frame,text="Удалить",command=delete).grid(row=1,column=10,padx=3)

    tree=ttk.Treeview(win,columns=labels,show="headings")
    for l in labels: tree.heading(l,text=l)
    tree.pack(expand=True,fill="both")

    load()

# ================== ПРЕПАРАТЫ ==================

def products_window():
    win=tk.Toplevel()
    win.title("Препараты")
    win.geometry("900x400")

    frame=tk.Frame(win);frame.pack(pady=5)

    labels = ["Код","Бренд","Серия","Срок годности"]
    entries=[]

    for i,l in enumerate(labels):
        tk.Label(frame,text=l).grid(row=0,column=i)
        e=tk.Entry(frame,width=18); e.grid(row=1,column=i)
        entries.append(e)

    def load():
        tree.delete(*tree.get_children())
        conn=get_connection();cur=conn.cursor()
        cur.execute("SELECT product_code,brand,batch,expiry_date FROM products")
        for r in cur.fetchall():
            tree.insert("", "end", values=r)
        conn.close()

    def add():
        conn=get_connection();cur=conn.cursor()
        cur.execute("""
            INSERT INTO products (product_code,brand,batch,expiry_date)
            VALUES (?,?,?,?)
        """, tuple(e.get() for e in entries))
        conn.commit();conn.close();load()

    def delete():
        sel=tree.selection()
        if not sel: return
        code=tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Удалить",f"Удалить препарат {code}?"):
            conn=get_connection();cur=conn.cursor()
            cur.execute("DELETE FROM products WHERE product_code=?", (code,))
            conn.commit();conn.close();load()

    def import_xlsx():
        p=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if not p: return
        df=pd.read_excel(p)
        conn=get_connection();cur=conn.cursor()
        for _,r in df.iterrows():
            cur.execute("""
                INSERT INTO products (product_code,brand,batch,expiry_date)
                VALUES (?,?,?,?)
            """, (r.get("product_code"),r.get("brand"),
                  r.get("batch"),r.get("expiry_date")))
        conn.commit();conn.close();load()

    def export_xlsx():
        conn=get_connection();cur=conn.cursor()
        cur.execute("SELECT product_code,brand,batch,expiry_date FROM products")
        rows=cur.fetchall();conn.close()
        os.makedirs("exports",exist_ok=True)
        wb=Workbook();ws=wb.active
        ws.append(["product_code","brand","batch","expiry_date"])
        for r in rows: ws.append(r)
        wb.save(f"exports/products_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

    tk.Button(frame,text="Добавить",command=add).grid(row=1,column=4,padx=3)
    tk.Button(frame,text="Импорт",command=import_xlsx).grid(row=1,column=5,padx=3)
    tk.Button(frame,text="Экспорт",command=export_xlsx).grid(row=1,column=6,padx=3)
    tk.Button(frame,text="Шаблон",
              command=lambda:download_template("products_template.xlsx")).grid(row=1,column=7,padx=3)
    tk.Button(frame,text="Удалить",command=delete).grid(row=1,column=8,padx=3)

    tree=ttk.Treeview(win,columns=labels,show="headings")
    for l in labels: tree.heading(l,text=l)
    tree.pack(expand=True,fill="both")

    load()

# ================== ПРОДАЖИ ==================

def sales_window():
    win=tk.Toplevel()
    win.title("Продажи")
    win.geometry("450x320")

    frame=tk.Frame(win);frame.pack(pady=10)

    tk.Label(frame,text="Врач").grid(row=0,column=0)
    tk.Label(frame,text="Препарат").grid(row=1,column=0)
    tk.Label(frame,text="Количество").grid(row=2,column=0)

    conn=get_connection();cur=conn.cursor()
    cur.execute("SELECT doctor_code,region FROM doctors")
    doc_map=dict(cur.fetchall())
    cur.execute("SELECT product_code FROM products")
    products=[r[0] for r in cur.fetchall()]
    conn.close()

    cb_doc=ttk.Combobox(frame,values=list(doc_map.keys()))
    cb_prod=ttk.Combobox(frame,values=products)
    e_qty=tk.Entry(frame)

    cb_doc.grid(row=0,column=1)
    cb_prod.grid(row=1,column=1)
    e_qty.grid(row=2,column=1)

    def save():
        conn=get_connection();cur=conn.cursor()
        cur.execute("""
            INSERT INTO sales (date,doctor_code,product_code,region,quantity)
            VALUES (?,?,?,?,?)
        """,(datetime.now().strftime("%Y-%m-%d"),
             cb_doc.get(),cb_prod.get(),
             doc_map.get(cb_doc.get(),""),int(e_qty.get())))
        conn.commit();conn.close()
        messagebox.showinfo("OK","Сохранено")

    tk.Button(win,text="Сохранить",command=save).pack(pady=5)
    tk.Button(win,text="Шаблон продаж",
              command=lambda:download_template("sales_template.xlsx")).pack(pady=5)

# ================== ОТЧЁТЫ ==================

# ================== СТАРТ ==================

if __name__ == "__main__":
    init_db()
    create_excel_templates()
    login_window()

